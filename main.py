import math
import multiprocessing
import random

import pygame
import pymsgbox
from openpyxl import Workbook

# Creating Microsoft Excel workbook for notating moves
wb = Workbook()
ws = wb.active
ws["A1"] = "White"
ws["B1"] = "Black"

if __name__ == "__main__":
    # Initialization and sound effects
    pygame.init()
    screen = pygame.display.set_mode((800, 800))
    icon = pygame.image.load("logo.png").convert_alpha()
    caption = pygame.display.set_caption("Chess Game")
    pygame.display.set_icon(icon)
    move = pygame.mixer.Sound("Sound Effects/Move.ogg")
    error_sound = pygame.mixer.Sound("Sound Effects/Error.ogg")
    victory = pygame.mixer.Sound("Sound Effects/Victory.ogg")
    checkmate_list = multiprocessing.Manager().list()

# Possible king and rook positions for castling
king_moves = [(7, 4), (0, 4)]
rook_moves = [(7, 7), (0, 0), (7, 0), (0, 7)]

# Initial Pawn Positions (ID LIST)
pawn_positions = [(1, 0), (1, 1), (1, 2), (1, 3), (1, 4), (1, 5), (1, 6), (1, 7), (6, 0), (6, 1), (6, 2), (6, 3),
                  (6, 4), (6, 5), (6, 6), (6, 7)]
# List for different pieces for calculating check
rooks, bishops, queens, knights, pawns, kings = [], [], [], [], [], []

# Linking each piece name string to it's corresponding list
pieces_to_piece_names = {"Bishop": bishops, "Rook": rooks, "Pawn": pawns, "Knight": knights, "Queen": queens,
                         "King": kings}

# Used songs through the gameplay
songs = ['Music/Chopin - Nocturne 1.mp3', 'Music/Chopin - Nocturne 2.mp3', 'Music/Schubert - Serenade.mp3',
         'Music/Tchaikovsky - Valse Sentimentale.mp3', 'Music/Beethoven Silence.mp3']

currently_playing_song = None
current_player, checked_player = "White", None
king_legal_moves_while_in_check = []
pieces_dictionary, click_count = {}, 0
reason_of_failure, running = None, 1
reject, pawn_passed = False, False
selected_piece, arrival_piece = None, None
blocking_piece, blocked_king = None, None
destination, alternative_destination = 0, 0
pieces_change, selected = [], 0
pieces, colored_rectangle, moved_pieces = [], [], []
en_passant, promoted_pawns, castling_rectangles = [], [], []
castling_locations, blocking_pieces = [2, 3, 5, 6], []
# Notation Management
notation_line, letter = 1, "A"
notation_columns = {0: "a", 1: "b", 2: "c", 3: "d", 4: "e", 5: "f", 6: "g", 7: "h"}
notation_rows = {0: 8, 1: 7, 2: 6, 3: 5, 4: 4, 5: 3, 6: 2, 7: 1}
notation_names = {"Knight": "N", "Queen": "Q", "King": "K", "Rook": "R", "Bishop": "B"}


def adding_notations(piece_name, original_file, capture_state, arrival_column, arrival_row):
    ws[letter + str(notation_line)] = piece_name + original_file + capture_state + arrival_column + arrival_row


# Basic Moves for king piece
def king_legal_moves(departure):
    the_legal_move = [(departure.id[0] - 1, departure.id[1] + 1), (departure.id[0] - 1, departure.id[1] - 1),
                      (departure.id[0] + 1, departure.id[1] + 1), (departure.id[0] + 1, departure.id[1] - 1),
                      (departure.id[0], departure.id[1] + 1), (departure.id[0], departure.id[1] - 1),
                      (departure.id[0] + 1, departure.id[1]), (departure.id[0] - 1, departure.id[1])]
    return the_legal_move


# Playing Shuffled Songs
if __name__ == "__main__":
    next_song = random.choice(songs)
    while next_song == currently_playing_song:
        next_song = random.choice(songs)
    currently_playing_song = next_song
    pygame.mixer.music.load(next_song)
    pygame.mixer.music.play()


# Getting the opposite piece color dynamically
def opposing_color(inputcolor):
    if inputcolor == "White":
        opposed_color = "Black"
    elif inputcolor == "Black":
        opposed_color = "White"
    else:
        opposed_color = None
    return opposed_color


# Conditions for the calculating check function to be executed
def conditions(start, end):
    global blocked_king, blocking_piece
    if end not in king_legal_moves_while_in_check:
        if opposing_color(start.color) == end.color and end.name == "King":
            return 1
        elif end.color is None and end.id[1] in castling_locations and (end.id[0] == 7 and start.color == "Black" or \
                                                                        end.id[
                                                                            0] == 0 and start.color == "White") and checked_player is None:
            return 2
    elif blocked_king is not None and opposing_color(
            start.color) == blocked_king.color and (end.color is None or end.color == start.color):
        return 3


# Checking if a certain piece blocks another piece (usually king)
def piece_check(iterable, arriving_piece):
    global blocked_king, blocking_piece
    if conditions(iterable, arriving_piece) == 1:  # The checked piece is a king
        blocking_piece = iterable
        blocked_king = arriving_piece
        if (blocking_piece.id, blocking_piece.color) not in blocking_pieces:
            blocking_pieces.append((blocking_piece.id, blocking_piece.color))
    elif conditions(iterable, arriving_piece) == 2:
        # The checked piece is in the king's path while castling
        if iterable.id not in castling_rectangles:
            castling_rectangles.append(iterable.id)
    elif conditions(iterable, arriving_piece) == 3:
        # The checked piece is one of the king's possible moves while in check
        if arriving_piece.id not in checkmate_list:
            checkmate_list.append(arriving_piece.id)


# Repeating the piece check function upon all the possible pieces
def calculating_check(beginning, arrival):
    global alternative_destination, blocking_piece, blocked_king, checked_player
    if beginning.name == "Bishop":
        for I in bishops:
            if possible_bishop(I, arrival) in ["True", alternative_destination]:
                piece_check(I, arrival)
    if beginning.name == "Queen":
        for I in queens:
            if possible_rook(I, arrival) is None:
                if possible_bishop(I, arrival) in ["True", alternative_destination]:
                    piece_check(I, arrival)
            elif possible_bishop(I, arrival) is None:
                if possible_rook(I, arrival) in ["True", alternative_destination]:
                    piece_check(I, arrival)
    if beginning.name == "Rook":
        for I in rooks:
            if possible_rook(I, arrival) in ["True", alternative_destination]:
                piece_check(I, arrival)
    if beginning.name == "Knight":
        for I in knights:
            if possible_knight(I, arrival) == alternative_destination:
                piece_check(I, arrival)
    if beginning.name == "Pawn":
        for I in pawns:
            if possible_pawn(I, arrival) == alternative_destination:
                piece_check(I, arrival)
    alternative_destination = 0
    if blocking_pieces == [] and len(castling_rectangles) == 0:  # There are no pieces which are blocked
        # Clearing the unnecessary variables
        king_legal_moves_while_in_check.clear()
        for x in checkmate_list:
            try:
                checkmate_list.remove(x)
            except ValueError:
                pass
        blocked_king = None
        blocking_piece = None
        checked_player = None


# Adding each piece from the pieces dictionary to it's convenient array (rooks, queens, pawns)
def minimal_adding(piece_name):
    if pieces_dictionary[j].name == piece_name:
        previous_color = [x.color for x in pieces_to_piece_names[piece_name] if x == pieces_dictionary[j]]
        for l in pieces_to_piece_names[piece_name]:
            if l in pieces_change:
                pieces_to_piece_names[piece_name].remove(l)
        if pieces_dictionary[j] not in pieces_to_piece_names[piece_name] or previous_color[0] != pieces_dictionary[
            j].color and len(previous_color) == 1:
            pieces_to_piece_names[piece_name].append(pieces_dictionary[j])


# Repeating the minimal adding function over all the pieces
def adding_pieces():
    for piece in pieces_to_piece_names:
        minimal_adding(piece)


# Function that takes care of the possible movements of the pawn
def possible_pawn(departing_point, arriving_point):
    going_ahead = False
    if logical_pawn(departing_point, arriving_point) == "True" and not hasattr(arriving_point, "piece"):
        if departing_point.id in pawn_positions:  # The pawn is in it's initial position
            for z in range(1, 3):  # The pawn can move 2 blocks ahead
                if departing_point.color == "White":
                    if arriving_point.id in [(departing_point.id[0] - z, departing_point.id[1])]:
                        going_ahead = True
                else:
                    if arriving_point.id in [(departing_point.id[0] + z, departing_point.id[1])]:
                        going_ahead = True
        else:  # The pawn has been already moved
            if departing_point.color == "White":
                r = -1
            else:
                r = 1
            # The pawn can only move 1 rectangle ahead
            if arriving_point.id in [(departing_point.id[0] + r, departing_point.id[1])]:
                going_ahead = True
        if arriving_point.id[0] in [departing_point.id[0] - 2, departing_point.id[0] + 2]:
            for r in pawns:
                if r.id in [(arriving_point.id[0], arriving_point.id[1] - 1),
                            (arriving_point.id[0], arriving_point.id[1] + 1)] and r.color == opposing_color(
                    departing_point.color) and going_ahead:
                    # En passant is permitted
                    if opposing_color(departing_point.color) == "Black":
                        z = 1
                    else:
                        z = -1
                    possible_move = [pieces_dictionary[x] for x in pieces_dictionary if
                                     pieces_dictionary[x].id in [(arriving_point.id[0] + z, arriving_point.id[1])]]
                    if departing_point not in en_passant:
                        en_passant.append(departing_point)
                    if possible_move[0] not in en_passant:
                        en_passant.append(possible_move[0])
                    if arriving_point not in en_passant:
                        en_passant.append(arriving_point)
        if going_ahead is True:
            # The pawn simply advances
            return "True"
    elif logical_pawn(departing_point, arriving_point) == alternative_destination:
        # The pawn captures a piece or puts it in check
        return alternative_destination


# Function that takes care of the possible movements of the knight
def possible_knight(departing_point, arriving_point):
    if arriving_point.id in [(departing_point.id[0] - 1, departing_point.id[1] + 2),
                             (departing_point.id[0] - 1, departing_point.id[1] - 2),
                             (departing_point.id[0] + 1, departing_point.id[1] + 2),
                             (departing_point.id[0] + 1, departing_point.id[1] - 2),
                             (departing_point.id[0] - 2, departing_point.id[1] + 1),
                             (departing_point.id[0] - 2, departing_point.id[1] - 1),
                             (departing_point.id[0] + 2, departing_point.id[1] + 1),
                             (departing_point.id[0] + 2, departing_point.id[1] - 1)]:
        if logical_k(arriving_point) == "True":
            return "True"
        if logical_k(arriving_point) == alternative_destination:
            return alternative_destination


# Function that takes care of the possible movements of the rook
def possible_rook(departing_point, arriving_point):
    # Rook is in the top
    if arriving_point.id[0] > departing_point.id[0] and arriving_point.id[1] == departing_point.id[1]:
        if logical_rook(departing_point.id[0], arriving_point.id[0], 0, 1, departing_point, arriving_point) == "True":
            return "True"
        elif logical_rook(departing_point.id[0], arriving_point.id[0], 0, 1, departing_point,
                          arriving_point) == alternative_destination:
            return alternative_destination
    # Rook is in the bottom
    elif arriving_point.id[0] < departing_point.id[0] and arriving_point.id[1] == departing_point.id[1]:
        if logical_rook(arriving_point.id[0], departing_point.id[0], 0, 1, departing_point, arriving_point) == "True":
            return "True"
        elif logical_rook(arriving_point.id[0], departing_point.id[0], 0, 1, departing_point,
                          arriving_point) == alternative_destination:
            return alternative_destination
    # Rook is on the right
    elif arriving_point.id[1] > departing_point.id[1] and arriving_point.id[0] == departing_point.id[0]:
        if logical_rook(departing_point.id[1], arriving_point.id[1], 1, 0, departing_point, arriving_point) == "True":
            return "True"
        elif logical_rook(departing_point.id[1], arriving_point.id[1], 1, 0, departing_point,
                          arriving_point) == alternative_destination:
            return alternative_destination
    # Rook is on the left
    elif arriving_point.id[1] < departing_point.id[1] and arriving_point.id[0] == departing_point.id[0]:
        if logical_rook(arriving_point.id[1], departing_point.id[1], 1, 0, departing_point, arriving_point) == "True":
            return "True"
        elif logical_rook(arriving_point.id[1], departing_point.id[1], 1, 0, departing_point,
                          arriving_point) == alternative_destination:
            return alternative_destination


# Function that takes care of the possible movements of the bishop
def possible_bishop(departing_point, arriving_point):
    if basic_bishop(arriving_point, departing_point):
        if arriving_point.id[1] > departing_point.id[1]:
            # The bishop and the arriving piece are on the primary diagonal
            if arriving_point.id[0] > departing_point.id[0]:  # The bishop is in the top
                if logical_bishop(departing_point.id[0], arriving_point.id[0], departing_point.id[1],
                                  arriving_point.id[1], departing_point, arriving_point) == "True":
                    return "True"
                elif logical_bishop(departing_point.id[0], arriving_point.id[0], departing_point.id[1],
                                    arriving_point.id[1], departing_point, arriving_point) == alternative_destination:
                    return alternative_destination
            elif arriving_point.id[0] < departing_point.id[0]:  # The bishop is in the bottom
                if logical_bishop(arriving_point.id[0], departing_point.id[0], departing_point.id[1],
                                  arriving_point.id[1], departing_point, arriving_point) == "True":
                    return "True"
                elif logical_bishop(arriving_point.id[0], departing_point.id[0], departing_point.id[1],
                                    arriving_point.id[1], departing_point, arriving_point) == alternative_destination:
                    return alternative_destination
        elif arriving_point.id[1] < departing_point.id[1]:
            # The bishop and the arriving piece are on the secondary diagonal
            if arriving_point.id[0] > departing_point.id[0]:  # The bishop is in the top
                if logical_bishop(departing_point.id[0], arriving_point.id[0], arriving_point.id[1],
                                  departing_point.id[1], departing_point, arriving_point) == "True":
                    return "True"
                elif logical_bishop(departing_point.id[0], arriving_point.id[0], arriving_point.id[1],
                                    departing_point.id[1], departing_point, arriving_point) == alternative_destination:
                    return alternative_destination
            elif arriving_point.id[0] < departing_point.id[0]:  # The bishop is in the bottom
                if logical_bishop(arriving_point.id[0], departing_point.id[0], arriving_point.id[1],
                                  departing_point.id[1], departing_point, arriving_point) == "True":
                    return "True"
                elif logical_bishop(arriving_point.id[0], departing_point.id[0], arriving_point.id[1],
                                    departing_point.id[1], departing_point, arriving_point) == alternative_destination:
                    return alternative_destination


# Basic Moves for bishop piece
def basic_bishop(arrival, departure):
    for b in range(1, 8):
        if arrival.id in [(departure.id[0] + b, departure.id[1] + b),
                          (departure.id[0] + b, departure.id[1] - b),
                          (departure.id[0] - b, departure.id[1] + b),
                          (departure.id[0] - b, departure.id[1] - b)]:
            return True


# Custom rectangle class that inherits from the original one
class Rectangle(pygame.Rect):
    def __init__(self, my_id, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.id = my_id
        self.name = None
        self.color = None
        # Assigning proper colors to each rectangle
        if self.id[0] in [0, 2, 4, 6, 8]:
            if self.id[1] in [0, 2, 4, 6, 8]:
                self.rect_color = (255, 255, 255)
            else:
                self.rect_color = (51, 102, 0)
        else:
            if self.id[1] in [0, 2, 4, 6, 8]:
                self.rect_color = (51, 102, 0)
            else:
                self.rect_color = (255, 255, 255)

    # Function for drawing pieces on the rectangle
    def draw(self):
        piece_name = pygame.image.load(f"Images/{self.color} {self.name}.png").convert_alpha()
        self.piece = screen.blit(piece_name, (self.centerx - 30, self.centery - 30))


# Calculating distance between two objects
def distance(x1, x2, y1, y2, distances):
    distance = math.sqrt((math.pow(x1 - x2, 2) + (math.pow(y1 - y2, 2))))
    distances.append([key, distance])


# Managing the bishop's movements
def logical_bishop(w, x, y, z, departure, arrival):
    global alternative_destination
    count = 0
    id_list = []
    distances = []
    distance(departure.centerx, arrival.centerx, departure.centery, arrival.centery, distances)
    if distances[0][1] > 142:  # If the Bishop is moving over multiple rectangles
        for l in pieces_dictionary:
            if basic_bishop(pieces_dictionary[l], departure):
                for J in range(w + 1, x):
                    for k in range(y + 1, z):
                        id_list.append((J, k))
                        if pieces_dictionary[l].id in id_list:
                            if pieces_dictionary[l].color is not None:
                                if arrival in king_legal_moves_while_in_check and pieces_dictionary[l] == blocked_king:
                                    alternative_destination = arrival
                                    count = 1
                                else:
                                    return "False"  # There is a piece in the path
                            if hasattr(arrival, "piece"):
                                alternative_destination = arrival
                                count = 1
        if count == 0:
            return "True"  # Bishop does a normal move
        else:
            return alternative_destination  # Bishop captures another piece
    else:  # If the rook is moving over 1 rectangle
        if hasattr(arrival, "piece") or arrival in king_legal_moves_while_in_check:
            alternative_destination = arrival
            return alternative_destination  # Bishop captures another piece
        else:
            return "True"  # Bishop does a normal move


# Managing the rook's movements
def logical_rook(w, x, y, z, departure, arrival):
    global alternative_destination
    count = 0
    distances = []
    distance(departure.centerx, arrival.centerx, departure.centery, arrival.centery, distances)
    if distances[0][1] > 100.0:  # If the rook is moving over multiple rectangles
        for t in range(w + 1, x):
            for m in pieces_dictionary:
                if pieces_dictionary[m].id[y] == t and pieces_dictionary[m].id[z] == departure.id[z]:
                    if pieces_dictionary[m].color is not None:
                        if arrival in king_legal_moves_while_in_check and pieces_dictionary[m] == blocked_king:
                            alternative_destination = arrival
                            count = 1
                        else:
                            return "False"  # There is a piece in the path
                    if hasattr(arrival, "piece"):
                        alternative_destination = arrival
                        count = 1
        if count == 0:
            return "True"  # Rook does a normal move
        else:
            return alternative_destination  # Rook captures another piece
    else:  # If the rook is moving over 1 rectangle
        if hasattr(arrival, "piece") or arrival in king_legal_moves_while_in_check:
            alternative_destination = arrival
            return alternative_destination  # Rook captures another piece
        else:
            return "True"  # Rook does a normal move


# Managing the knight and king's movements
def logical_k(arrival):
    global alternative_destination
    for t in pieces_dictionary:
        if pieces_dictionary[t] == arrival:
            if hasattr(arrival, "piece"):
                alternative_destination = arrival
                return alternative_destination  # The arrival rectangle has a piece
            else:
                return "True"  # The arrival rectangle is empty


# Managing the pawn's movements
def logical_pawn(departure, arrival):
    global pawn_passed, alternative_destination
    if opposing_color(departure.color) == "Black":
        e = -1
    else:
        e = 1
    for r in pieces_dictionary:
        if pieces_dictionary[r] == arrival:
            if arrival.id[0] == departure.id[0] + e and arrival.id[1] in [departure.id[1] + 1, departure.id[1] - 1]:
                alternative_destination = arrival
                if en_passant != [] and not hasattr(arrival, "piece"):  # The pawn did an en passant
                    pieces_change.append(en_passant[2])
                    moved_pieces.append(en_passant[2])
                    pawn_passed = True
                return alternative_destination  # The pawn captures or puts another piece in check
            else:
                return "True"  # The pawn did a simple move


def castling(chosen_piece, second_piece, notation):
    global notation_line, letter
    if len(castling_rectangles) == 0:
        if current_player == "White":
            letter = "A"
            notation_line += 1
        else:
            letter = "B"
        moved_pieces.append(chosen_piece)
        pieces.append(
            [second_piece, chosen_piece.piece,
             chosen_piece.name,
             chosen_piece.color])
        if second_piece in pieces_change:
            pieces_change.remove(second_piece)
        if chosen_piece.id in rook_moves:
            rook_moves.remove(chosen_piece.id)
        ws[letter + str(notation_line)] = notation
        return True


def worker(checkmate_list):
    for s in king_legal_moves_while_in_check:
        for g in pieces_list:
            if conditions(g, s) == 3:
                calculating_check(g, s)


if __name__ == "__main__":
    # Main loop
    while running:
        # Multiprocessing for checking checkmate
        p = multiprocessing.Process(target=worker, args=[checkmate_list])
        p.start()
        pieces_list = rooks + bishops + pawns + knights + queens
        for s in kings:  # Calculating check in real time
            for g in pieces_list:
                if conditions(g, s) == 1:
                    calculating_check(g, s)
        for s in pieces_dictionary:  # Checking if castling is allowed
            for g in pieces_list:
                if conditions(g, pieces_dictionary[s]) == 2:
                    calculating_check(g, pieces_dictionary[s])
        for event in pygame.event.get():
            # Looping through events
            if event.type == pygame.QUIT:
                the_box = pymsgbox.confirm("Are you sure you want to quit?", "Are you sure?")
                if the_box == "OK":
                    if notation_line == 1:
                        notation_line += 1
                    ws[letter + str(notation_line)] = "1/2"
                    wb.save("Logs/Game_Details.xlsx")
                    running = 0
            if event.type == pygame.MOUSEBUTTONDOWN and event.button == pygame.BUTTON_LEFT:
                # Left Mouse Button Pressed
                click_count += 1
                pos = pygame.mouse.get_pos()
                distances1 = []
                for key in pieces_dictionary:
                    distance(key[0], pos[0], key[1], pos[1], distances1)
                smallest_distance = distances1[0][1]
                for i in range(len(distances1)):
                    if distances1[i][1] < smallest_distance:
                        smallest_distance = distances1[i][1]
                coordinations = [x for x in distances1 if smallest_distance in x]
                rectangle = pieces_dictionary[coordinations[0][0]]
                if click_count % 2 == 1:
                    # On player first click
                    if hasattr(rectangle, "piece"):
                        if current_player == "White":
                            if rectangle.color == "White":
                                selected = rectangle
                                colored_rectangle = selected
                            else:
                                # Hide window until alert dismissed
                                pygame.display.set_mode((1, 1))
                                error_sound.play()
                                the_box = pymsgbox.alert("It's White's Turn", "Wrong Choice")
                                if the_box == "OK":
                                    # Show window again
                                    pygame.display.set_mode((800, 800))
                                click_count -= 1
                        elif current_player == "Black":
                            if rectangle.color == "Black":
                                selected = rectangle
                                colored_rectangle = selected
                            else:
                                # Hide window until alert dismissed
                                pygame.display.set_mode((1, 1))
                                error_sound.play()
                                the_box = pymsgbox.alert("It's Black's Turn", "Wrong Choice")
                                if the_box == "OK":
                                    # Show window again
                                    pygame.display.set_mode((800, 800))
                                click_count -= 1
                    else:
                        # Hide window until alert dismissed
                        pygame.display.set_mode((1, 1))
                        error_sound.play()
                        the_box = pymsgbox.alert("Empty Box Chosen", "Wrong Choice")
                        if the_box == "OK":
                            # Show window again
                            pygame.display.set_mode((800, 800))
                        click_count -= 1
                elif click_count % 2 == 0:
                    # On player's second click
                    if hasattr(rectangle, "piece") and rectangle.color == selected.color:
                        # Hide window until alert dismissed
                        pygame.display.set_mode((1, 1))
                        error_sound.play()
                        the_box = pymsgbox.alert("Choose a correct position", "Wrong Choice")
                        if the_box == "OK":
                            # Show window again
                            pygame.display.set_mode((800, 800))
                        colored_rectangle = []
                    else:
                        destination = rectangle
                        going_ahead = False
                        go_king = False
                        pawn_promotion = False
                        # Piece Letter for notations
                        piece_letter = ""
                        if selected.name == "Pawn":
                            # Controlling Pawn Moves
                            if len(blocking_pieces) == 0 or current_player != checked_player:
                                if possible_pawn(selected, destination) == alternative_destination and (
                                        destination.color == opposing_color(selected.color) or en_passant != []):
                                    destination = alternative_destination
                                    going_ahead = True
                                elif possible_pawn(selected, destination) == "True":
                                    going_ahead = True
                                else:
                                    reason_of_failure = "Incorrect Position"
                                if (selected.color == "White" and destination.id[0] == 0) or (
                                        selected.color == "Black" and destination.id[0] == 7):
                                    # Hide window until alert dismissed
                                    pygame.display.set_mode((1, 1))
                                    the_box = pymsgbox.alert("The pawn will be promoted to a Queen!", "Great!")
                                    if the_box == "OK":
                                        # Show window again
                                        pygame.display.set_mode((800, 800))
                                        promoted_pawns.append(destination)
                                    pawn_promotion = True
                            else:
                                reason_of_failure = "Player king is in check!"
                        if selected.name == "Rook" or selected.name == "Queen":
                            # Controlling Rook moves (Queen can move like Rook)
                            if len(blocking_pieces) == 0 or current_player != checked_player:
                                if possible_rook(selected, destination) == "True":
                                    going_ahead = True
                                elif possible_rook(selected, destination) == alternative_destination:
                                    destination = alternative_destination
                                    going_ahead = True
                                else:
                                    reason_of_failure = "Incorrect Position"
                            else:
                                reason_of_failure = "Player king is in check!"
                            piece_letter = selected.name[0]
                        if selected.name == "Bishop" or selected.name == "Queen":
                            # Controlling Bishop moves (Queen can move like Bishop)
                            if len(blocking_pieces) == 0 or current_player != checked_player:
                                if possible_bishop(selected, destination) == alternative_destination:
                                    destination = alternative_destination
                                    going_ahead = True
                                elif possible_bishop(selected, destination) == "True":
                                    going_ahead = True
                                else:
                                    reason_of_failure = "Incorrect Position"
                            else:
                                reason_of_failure = "Player king is in check!"
                            piece_letter = selected.name[0]
                        if selected.name == "Knight":
                            # Controlling Knight moves
                            if len(blocking_pieces) == 0 or current_player != checked_player:
                                if possible_knight(selected, destination) == alternative_destination:
                                    going_ahead = True
                                    destination = alternative_destination
                                elif possible_knight(selected, destination) == "True":
                                    going_ahead = True
                                else:
                                    reason_of_failure = "Incorrect Position"
                            else:
                                reason_of_failure = "Player king is in check!"
                            piece_letter = "N"
                        if selected.name == "King":
                            # Controlling King Moves
                            if current_player == checked_player or len(blocking_pieces) == 0:
                                if destination.id in king_legal_moves(selected) and destination.name != "King":
                                    # King does a normal move
                                    for i in checkmate_list:
                                        checkmate_list.remove(i)
                                    blocking_pieces.clear()
                                    if logical_k(destination) == "True":
                                        going_ahead = True
                                        go_king = True
                                    elif logical_k(destination) == alternative_destination:
                                        destination = alternative_destination
                                        going_ahead = True
                                        go_king = True
                                    else:
                                        reason_of_failure = "King will be put in check or is already in check"
                                elif not hasattr(destination, "piece") and (
                                        len(blocking_pieces) == 0 or checked_player != current_player):
                                    # King attempts castling
                                    if selected.color == "White":
                                        line = 7
                                    else:
                                        line = 0
                                    for i in pieces_dictionary:
                                        if pieces_dictionary[i].id == (line, destination.id[1] - 1):
                                            piece_1 = pieces_dictionary[i]
                                        elif pieces_dictionary[i].id == (line, destination.id[1] - 2):
                                            piece_2 = pieces_dictionary[i]
                                        elif pieces_dictionary[i].id == (line, destination.id[1] + 1):
                                            piece_3 = pieces_dictionary[i]
                                    if not hasattr(piece_1, "piece"):
                                        if selected.id in king_moves:
                                            if destination.id == (line, 6) and piece_3.id in rook_moves:
                                                # Castling King Side
                                                if castling(piece_3, piece_1, "0-0"):
                                                    going_ahead = True
                                                else:
                                                    reason_of_failure = "King will fall or move through check"
                                            elif destination.id == (line, 2) and not hasattr(piece_3,
                                                                                             "piece") and piece_2.id in rook_moves:
                                                # Castling Queen Side
                                                if castling(piece_2, piece_3, "0-0-0"):
                                                    going_ahead = True
                                                else:
                                                    reason_of_failure = "King will fall or move through check"
                            if not going_ahead and reason_of_failure is None:
                                reason_of_failure = "Incorrect Move"
                            piece_letter = "K"
                        if going_ahead:
                            # On successful move
                            if go_king and selected.name == "King" or selected.name != "King":
                                if current_player == "White":
                                    letter = "A"
                                    notation_line += 1
                                else:
                                    letter = "B"
                                if alternative_destination != 0:
                                    # Selected Piece captures another piece
                                    adding_notations(piece_letter, str(notation_columns[selected.id[1]]), "x",
                                                     str(notation_columns[destination.id[1]]),
                                                     str(notation_rows[destination.id[0]]))
                                else:
                                    # Selected Piece moves into an empty rectangle
                                    adding_notations(piece_letter, str(notation_columns[selected.id[1]]), "",
                                                     str(notation_columns[destination.id[1]]),
                                                     str(notation_rows[destination.id[0]]))
                                if pawn_promotion:
                                    # Selected piece (Pawn) is promoted (Automatically to Queen)
                                    ws[letter + str(notation_line)] = ws[letter + str(notation_line)].value + "=Q"
                            move.play()
                            # Changing current_player value to next player's color
                            current_player = opposing_color(current_player)
                            if selected not in pieces_change:
                                # Tracking pieces outside of the rectangle specified in the pieces array
                                pieces_change.append(selected)
                            if selected not in moved_pieces:
                                # Tracking pieces outside their original rectangle
                                moved_pieces.append(selected)
                            if alternative_destination not in pieces_change:
                                # Tracking pieces outside of the rectangle specified in the pieces array
                                pieces_change.append(alternative_destination)
                            if alternative_destination not in moved_pieces:
                                # Tracking pieces outside their original rectangle
                                moved_pieces.append(alternative_destination)
                            if selected.id in king_moves:
                                # Removing king piece ID from the king ID list to prohibit castling
                                king_moves.remove(selected.id)
                            if selected.id in rook_moves:
                                # Removing rook piece ID from the rook ID list to prohibit castling
                                rook_moves.remove(selected.id)
                            castling_rectangles.clear()
                        else:
                            error_sound.play()
                            # Hide windows until alert dismissed
                            pygame.display.set_mode((1, 1))
                            the_box = pymsgbox.alert(reason_of_failure, "Wrong Choice")
                            reason_of_failure = None
                            if the_box == "OK":
                                # Show window again
                                pygame.display.set_mode((800, 800))
                            destination = 0
                        colored_rectangle = []
                        alternative_destination = 0
            if event.type == pygame.KEYDOWN and event.key == pygame.K_ESCAPE:
                # Undo selecting piece
                if click_count % 2 == 1:
                    colored_rectangle = []
                    click_count -= 1
        # Filling screen with 64 rectangles
        for j in range(8):
            for i in range(8):
                # Initialising the rectangle and giving him and ID (J is the line and I is the column)
                rectangle = Rectangle((j, i), int(f"{i}00"), int(f"{j}00"), 100, 100)
                if not colored_rectangle:
                    screen.fill(rectangle.rect_color, rectangle)
                elif colored_rectangle == rectangle:
                    # Drawing Rectangle in red while clicked
                    screen.fill((255, 0, 0), rectangle)
                # Mapping each rectangle with it's coordinates inside a dictionary
                pieces_dictionary[(rectangle.centerx, rectangle.centery)] = rectangle
        for j in pieces_dictionary:
            if pieces_dictionary[j] in pieces_change:  # Removing moved pieces from the pieces array
                for k in pieces:
                    if pieces_dictionary[j] in k:
                        pieces.remove(k)
            key = pieces_dictionary[j].id
            for i in range(8):
                if pieces_dictionary[j] not in moved_pieces:  # Avoiding drawing already moved pieces
                    # Drawing each piece in it's initial position
                    if key[0] == 0 or key[0] == 1:
                        pieces_dictionary[j].color = "Black"
                    if key[0] == 6 or key[0] == 7:
                        pieces_dictionary[j].color = "White"
                    if key[0] in [1, 6]:
                        pieces_dictionary[j].name = "Pawn"
                        pieces_dictionary[j].draw()
                    elif key[0] in [0, 7]:
                        if key[1] in [0, 7]:
                            pieces_dictionary[j].name = "Rook"
                            pieces_dictionary[j].draw()
                        elif key[1] in [1, 6]:
                            pieces_dictionary[j].name = "Knight"
                            pieces_dictionary[j].draw()
                        elif key[1] in [2, 5]:
                            pieces_dictionary[j].name = "Bishop"
                            pieces_dictionary[j].draw()
                        elif key[1] == 3:
                            pieces_dictionary[j].name = "Queen"
                            pieces_dictionary[j].draw()
                        elif key[1] == 4:
                            pieces_dictionary[j].name = "King"
                            pieces_dictionary[j].draw()
                    adding_pieces()
            if destination == pieces_dictionary[j]:
                # Pieces array takes care of re-drawing moved pieces
                pieces.append([pieces_dictionary[j], selected.piece, selected.name, selected.color])
                if pieces_dictionary[j] in pieces_change:
                    pieces_change.remove(pieces_dictionary[j])
                if pieces_dictionary[j] in promoted_pawns and selected.name != "Pawn":
                    # Making sure pieces which capture promoted pawns aren't transformed into Queens
                    promoted_pawns.remove(pieces_dictionary[j])
                if pieces_dictionary[j] not in moved_pieces:
                    # Making sure that the script won't draw the default piece on the destination rectangle
                    moved_pieces.append(pieces_dictionary[j])
                if pawn_passed:
                    # Removing pawn from rectangle after en passant
                    if en_passant[2] in pieces_change:
                        pieces_change.remove(en_passant[2])
                    if en_passant[2] in moved_pieces:
                        moved_pieces.remove(en_passant[2])
                    en_passant.clear()
                    pawn_passed = False
                elif en_passant != [] and selected.color == opposing_color(en_passant[0].color):
                    # Protecting against doing en passant if not done the move when it was permitted
                    en_passant.clear()
                    pawn_passed = False
                selected_piece = selected
                arrival_piece = destination
                destination = 0
            for i in pieces:
                # Assigning new data to rectangle that was filled after piece movement
                if pieces_dictionary[j] in i:
                    pieces_dictionary[j].piece = i[1]
                    if pieces_dictionary[j] in promoted_pawns:
                        pieces_dictionary[j].name = "Queen"
                    else:
                        pieces_dictionary[j].name = i[2]
                    pieces_dictionary[j].color = i[3]
                    adding_pieces()
                    pieces_dictionary[j].draw()
                if selected_piece is not None and reject and arrival_piece in i:
                    # Protecting against putting own king in check
                    pieces.remove(i)
                    if selected_piece in pieces_change:
                        pieces_change.remove(selected_piece)
                    if selected_piece.name == "Pawn":
                        pawns.remove(arrival_piece)
                    elif selected_piece.name == "Knight":
                        knights.remove(arrival_piece)
                    elif selected_piece.name == "King":
                        kings.remove(arrival_piece)
                    pieces.append([selected_piece, selected_piece.piece, selected_piece.name, selected_piece.color])
                    if hasattr(arrival_piece, "piece"):
                        pieces.append([arrival_piece, arrival_piece.piece, arrival_piece.name, arrival_piece.color])
                    error_sound.play()
                    # Hide windows until alert dismissed
                    pygame.display.set_mode((1, 1))
                    the_box = pymsgbox.alert("King will be put in check or is already in check", "Wrong Choice")
                    if the_box == "OK":
                        # Show window again
                        pygame.display.set_mode((800, 800))
                    current_player = opposing_color(current_player)
                    if selected_piece.color == "White":
                        notation_line -= 1
                    else:
                        ws[letter + str(notation_line)] = ""
                    selected_piece = None
                    arrival_piece = None
                    blocking_pieces.clear()
                    reject = False
        if selected_piece is not None:
            if len(blocking_pieces) != 0 and selected_piece.color == checked_player:
                reject = True
            else:
                if blocked_king is not None:  # Calculating possible king moves while in check
                    count = 0
                    for K in blocking_pieces:
                        if K[1] == opposing_color(blocked_king.color):
                            count += 1
                    if count == len(blocking_pieces):
                        for L in pieces_dictionary:
                            for I in king_legal_moves(blocked_king):
                                if pieces_dictionary[L].id == I and pieces_dictionary[L].color != blocked_king.color:
                                    if pieces_dictionary[L] not in king_legal_moves_while_in_check:
                                        king_legal_moves_while_in_check.append(pieces_dictionary[L])
                        if len(king_legal_moves_while_in_check) == len(checkmate_list):  # Checking for checkmate
                            if "+" in ws[letter + str(notation_line)].value:
                                new_move = ws[letter + str(notation_line)].value.replace("+", "#")
                                ws[letter + str(notation_line)] = new_move
                            else:
                                ws[letter + str(notation_line)] = ws[letter + str(notation_line)].value + "#"
                            # Changing check notation to checkmate notation
                            victory.play()
                            wb.save("Logs/Game_Details.xlsx")
                            pymsgbox.alert(f"{blocking_piece.color} has won the game!", "Congratulations")
                            running = 0
                    elif count != len(blocking_pieces):
                        reject = True
        # Checking if the king is blocked
        if len(blocking_pieces) != 0:
            checked_player = opposing_color(blocking_piece.color)
            # Adding check notations
            if "+" not in ws[letter + str(notation_line)].value:
                ws[letter + str(notation_line)] = ws[letter + str(notation_line)].value + "+"
        # Updating the display
        pygame.display.update()
